import csv
import base64
from io import StringIO, BytesIO
from openpyxl import load_workbook
from odoo.exceptions import ValidationError
from odoo import api, fields, models, Command, _


class ImportSaleLine(models.TransientModel):
    _name = 'import.sale.line'
    _rec_name = 'select_file'
    _description = 'Import Sale Line'

    order_id = fields.Many2one(comodel_name="sale.order", string="Order")
    select_file = fields.Binary(string='Select File')
    file_name = fields.Char(string="File Name")
    import_by = fields.Selection(string="Import By", required=True, selection=[
        ('name', 'Name'), ('default_code', 'Default Code'), ('barcode', 'Barcode')], default='default_code')

    def action_import(self):
        imp_file = BytesIO(base64.b64decode(self.select_file))
        if self.file_name.endswith('.csv'):
            try:
                csv_file = StringIO(imp_file.read().decode())
                file_lines = csv.DictReader(csv_file, delimiter=',')
            except Exception as e:
                raise ValidationError(_(str(e)))
        elif self.file_name.endswith('.xlsx'):
            try:
                workbook = load_workbook(filename=imp_file)
                worksheet = workbook.active
                column_header = self.read_xlsx_file_header(worksheet)
                file_lines = self.get_xlsx_file_lines(worksheet, column_header)
            except Exception as e:
                raise ValidationError(_(str(e)))
        else:
            raise ValidationError(_('Please provide only .xlsx , .csv file'))

        if file_lines:
            self._prepare_order_lines(file_lines)

    def _prepare_order_lines(self, file_lines):
        product_obj = self.env['product.product']
        for line in file_lines:
            product_id = False
            product_domain = []
            if self.import_by == 'name':
                if not line.get('Name'):
                    continue;
                product_domain.append(('name', '=', line.get('Name')))
                product_id = product_obj.search(product_domain, limit=1)
                if not product_id:
                    continue;
            elif self.import_by == 'default_code':
                if not line.get('Code'):
                    continue;
                product_domain.append(('default_code', '=', line.get('Code')))
                product_id = product_obj.search(product_domain, limit=1)
                if not product_id:
                    continue;
            else:
                if not line.get('Barcode'):
                    continue;
                product_domain.append(('barcode', '=', line.get('Barcode')))
                product_id = product_obj.search(product_domain, limit=1)
                if not product_id:
                    continue;
            order_line = self.order_id.order_line.filtered(lambda l: l.product_id == product_id)
            product_uom_id = self.env['uom.uom'].search([('name', 'ilike', line.get('UOM'))], limit=1)
            if order_line:
                order_line.write({
                    "product_uom_qty": order_line.product_uom_qty + line.get('Quantity', 0),
                    'price_unit': line.get('Price', 0),
                    'product_uom': product_uom_id.id
                })
            else:
                new_line = [(0, 0, {
                    "product_id": product_id.id,
                    "product_uom_qty": line.get('Quantity', 0),
                    'price_unit': line.get('Price', 0),
                    'product_uom_id': product_uom_id.id
                })]
                self.order_id.update(dict(order_line=new_line))

    @staticmethod
    def read_xlsx_file_header(worksheet):
        try:
            column_header = {}
            columns = worksheet.max_column
            columns = columns + 1
            for row in worksheet.iter_rows():
                if row[0].row > 1 or not next((r for r in row if r.row), None):
                    break
                for col in range(1, columns):
                    column_header[col] = str(worksheet.cell(row=1, column=col).value)
            return column_header
        except Exception as e:
            raise ValidationError(str(e))

    @staticmethod
    def get_xlsx_file_lines(worksheet, column_header):
        try:
            data = []
            columns = worksheet.max_column
            columns = columns + 1
            for row_num, row in enumerate(worksheet.iter_rows()):
                row_num += 1
                if row[0].row == 1 or not next((r for r in row if r.row), None):
                    continue
                data.append({
                    column_header.get(col): worksheet.cell(row=row_num, column=col).value
                    for col in range(1, columns)
                })
            return data
        except Exception as e:
            raise ValidationError(_(str(e)))


